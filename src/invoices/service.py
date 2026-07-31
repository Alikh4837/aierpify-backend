# src\invoices\service.py
from __future__ import annotations

import math
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple, cast
from uuid import UUID

import xlrd
from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlmodel import col, select

from src.auth.user import AuthUser
from src.customers.enums import RegistrationTypeEnum
from src.customers.models import Customer
from src.customers.schemas import (
    CreateCustomerRequest,
    GetCustomerRequest,
)
from src.customers.service import CustomerService
from src.enums import ProvinceEnum, SaleTypeEnum
from src.exceptions import BadRequestException, NotFoundException
from src.invoices.enums import (
    InvoiceFBRStatusEnum,
    InvoiceStatusEnum,
    InvoiceTypeEnum,
)
from src.invoices.models import Invoice, InvoiceItem
from src.invoices.schemas import (
    CreateInvoiceItemResponse,
    CreateInvoiceItemsRequest,
    CreateInvoiceItemsResponse,
    CreateInvoiceRequest,
    CreateInvoiceResponse,
    DeleteInvoiceItemRequest,
    DeleteInvoiceItemResponse,
    DeleteInvoiceRequest,
    DeleteInvoiceResponse,
    ExportInvoicesExcelRequest,
    GetInvoiceCompleteResponse,
    GetInvoiceItemsRequest,
    GetInvoiceItemsResponse,
    GetInvoiceRequest,
    GetInvoiceResponse,
    GetInvoicesRequest,
    GetInvoicesResponse,
    InvoiceImportParsedInvoice,
    InvoiceImportParsedInvoiceItem,
    InvoiceImportParseResponse,
    InvoiceImportSubmitInvoice,
    InvoiceImportSubmitInvoiceItem,
    InvoiceImportSubmitInvoiceResult,
    InvoiceImportSubmitRequest,
    InvoiceImportSubmitResponse,
    InvoiceImportValidationError,
    InvoiceItemBase,
    InvoiceItemResponse,
    InvoiceResponse,
    UpdateInvoiceItemRequest,
    UpdateInvoiceItemResponse,
    UpdateInvoiceItemsRequest,
    UpdateInvoiceItemsResponse,
    UpdateInvoiceRequest,
    UpdateInvoiceResponse,
)
from src.products.models import Product
from src.products.schemas import CreateProductRequest
from src.products.service import ProductService
from src.users.schemas import (
    FBRProfileResponse,
    GetFBRProfileRequest,
    GetUserProfileRequest,
)
from src.users.service import FBRProfileService, UserPlanService, UserProfileService
from src.utils import apply_ordering_sql, apply_pagination_sql, get_user_id


# --------------------------------------------------------------------------- #
#                               CRUD Services                                 #
# --------------------------------------------------------------------------- #
class InvoiceService:
    @staticmethod
    async def get_invoice_single(
        auth_user: AuthUser, input_params: GetInvoiceRequest
    ) -> GetInvoiceResponse:
        """
        Fetch a single invoice by ID ensuring ownership.
        """

        session = auth_user.session
        user_id = get_user_id(auth_user, input_params.user_id)

        stmt = (
            select(Invoice)
            .where(Invoice.user_id == user_id)
            .where(Invoice.id == input_params.id)
        )
        response = await session.exec(stmt)
        data = response.first()

        if not data:
            raise NotFoundException(
                message="Invoice not found",
                detail=f"No invoice found with ID {input_params.id}",
                extra={"invoice_id": str(input_params.id)},
            )

        return GetInvoiceResponse.model_validate(data.model_dump())

    @staticmethod
    async def get_invoice_single_complete(
        auth_user: AuthUser,
        input_params: GetInvoiceRequest,
        include_tokens: bool = False,
    ) -> GetInvoiceCompleteResponse:
        """
        Fetch a single invoice by ID along with its items, ensuring ownership.
        """

        # Fetch invoice
        invoice = await InvoiceService.get_invoice_single(auth_user, input_params)
        if not invoice:
            raise NotFoundException(
                message="Invoice not found",
                detail=f"No invoice found with ID {input_params.id}",
                extra={"invoice_id": str(input_params.id)},
            )

        # Fetch associated user profile
        user_profile = await UserProfileService.get_user_profile(
            auth_user, input_params=GetUserProfileRequest(user_id=invoice.user_id)
        )
        if not user_profile:
            raise NotFoundException(
                message="User profile not found",
                detail=f"No user profile found for user ID {invoice.user_id}",
                extra={"user_id": str(invoice.user_id)},
            )

        # Fetch associated FBR profile
        fbr_profile_response = await FBRProfileService.get_fbr_profile(
            auth_user, GetFBRProfileRequest(user_id=invoice.user_id)
        )
        if not fbr_profile_response:
            raise NotFoundException(
                message="FBR profile not found",
                detail=f"No FBR profile found for user ID {invoice.user_id}",
                extra={"user_id": str(invoice.user_id)},
            )

        # Hide tokens if not requested
        if not include_tokens:
            fbr_profile_response.sandbox_token = ""
            fbr_profile_response.production_token = ""

        # Convert to public response model
        fbr_profile = FBRProfileResponse.model_validate(
            fbr_profile_response.model_dump()
        )

        # Fetch associated invoice items
        items = await InvoiceItemService.get_invoice_items(
            auth_user, GetInvoiceItemsRequest(invoice_id=invoice.id)
        )
        if not items:
            raise NotFoundException(
                message="Invoice items not found",
                detail=f"No invoice items found for invoice ID {invoice.id}",
                extra={"invoice_id": str(invoice.id)},
            )

        # Fetch associated products
        # Extract unique product IDs from invoice items
        product_ids: List[UUID] = list(
            set(item.product_id for item in items.data if item.product_id)
        )
        products = await ProductService.get_products_by_ids(auth_user, product_ids)

        # Fetch associated customer
        customer = await CustomerService.get_customer_single(
            auth_user,
            GetCustomerRequest(id=invoice.customer_id),
        )
        if not customer:
            raise NotFoundException(
                message="Customer not found",
                detail=f"No customer found with ID {invoice.customer_id}",
                extra={"customer_id": str(invoice.customer_id)},
            )

        return GetInvoiceCompleteResponse(
            **invoice.model_dump(),
            user_profile=user_profile,
            fbr_profile=fbr_profile,
            customer=customer,
            items=items.data,
            products=products,
        )

    @staticmethod
    async def get_invoices(
        auth_user: AuthUser, input_params: GetInvoicesRequest
    ) -> GetInvoicesResponse:
        """
        Retrieve invoices owned by the authenticated user with filtering & pagination.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_params: Filtering & pagination parameters.

        Returns:
            GetInvoicesResponse: Paginated invoice list.
        """

        session = auth_user.session
        user_id = get_user_id(auth_user, input_params.user_id)

        # Base ownership filter
        conditions: List[object] = [Invoice.user_id == user_id]

        # Check if ID filter is applied
        if input_params.id:
            conditions.append(Invoice.id == input_params.id)

        # If ID is provided, ignore other filters and fetch by ID only
        else:
            if input_params.customer_id:
                conditions.append(Invoice.customer_id == input_params.customer_id)

            if input_params.invoice_number:
                conditions.append(Invoice.invoice_number == input_params.invoice_number)

            if input_params.status:
                conditions.append(Invoice.status == input_params.status)

            if input_params.invoice_type:
                conditions.append(Invoice.invoice_type == input_params.invoice_type)

            if input_params.fbr_status:
                conditions.append(Invoice.fbr_status == input_params.fbr_status)

            if input_params.fbr_validated is not None:
                conditions.append(Invoice.fbr_validated == input_params.fbr_validated)

            if input_params.issue_date:
                conditions.append(Invoice.issue_date == input_params.issue_date)

            if input_params.due_date:
                conditions.append(Invoice.due_date == input_params.due_date)

        page = input_params.page
        page_size = input_params.page_size

        # Count total
        count_stmt = select(func.count()).select_from(Invoice).where(*conditions)  # type: ignore[arg-type]
        total_result = await session.exec(count_stmt)
        total = total_result.one()

        # Pagination
        stmt = (
            select(Invoice).where(*conditions)  # type: ignore[arg-type]
        )

        # Apply pagination
        stmt = await apply_pagination_sql(stmt, page, page_size)

        # Apply ordering
        stmt = await apply_ordering_sql(
            stmt, Invoice, input_params.order, input_params.order_by
        )

        result = await session.exec(stmt)
        rows: Sequence[Invoice] = result.all()

        # Cast/validate to response models
        data: List[InvoiceResponse] = [
            InvoiceResponse.model_validate(r.model_dump()) for r in rows
        ]

        next_page = total > page * page_size
        return GetInvoicesResponse(data=data, total=total, next_page=next_page)

    @staticmethod
    async def create_invoice(
        auth_user: AuthUser, input_data: CreateInvoiceRequest
    ) -> CreateInvoiceResponse:
        """
        Create a new invoice owned by the authenticated user.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_data: Invoice creation data.
        """

        session = auth_user.session
        user_id = get_user_id(auth_user, input_data.user_id)
        input_data.user_id = user_id

        # Check for duplicate invoice number if provided
        if input_data.invoice_number and input_data.invoice_number.strip():
            # Check for existing invoice with the same number for this user
            stmt = (
                select(Invoice)
                .where(Invoice.user_id == user_id)
                .where(Invoice.invoice_number == input_data.invoice_number)
            )
            result = await session.exec(stmt)
            existing_invoice = result.first()
            if existing_invoice:
                raise BadRequestException(
                    message="Invoice number already exists",
                    detail=f"An invoice with number {input_data.invoice_number} already exists.",
                    extra={"invoice_number": input_data.invoice_number},
                )

        payload = input_data.model_dump()

        # Queue the usage counter increment
        # This will be committed after the product is created
        await UserPlanService.increment_usage_counter(
            auth_user, counter_field="invoices_used", increment_by=1
        )

        invoice = Invoice(**payload)
        session.add(invoice)

        await session.commit()
        await session.refresh(invoice)

        return CreateInvoiceResponse.model_validate(invoice.model_dump())

    @staticmethod
    async def update_invoice(
        auth_user: AuthUser, input_data: UpdateInvoiceRequest
    ) -> UpdateInvoiceResponse:
        """
        Update an existing invoice owned by the authenticated user.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_data: Invoice update data.
        """

        session = auth_user.session
        user_id = get_user_id(auth_user, input_data.user_id)

        stmt = (
            select(Invoice)
            .where(Invoice.user_id == user_id)
            .where(Invoice.id == input_data.id)
        )
        result = await session.exec(stmt)
        invoice = result.first()

        if not invoice:
            raise NotFoundException(
                message="Invoice not found",
                detail=f"No invoice found with ID {input_data.id}",
                extra={"invoice_id": str(input_data.id)},
            )

        invoice_fields = Invoice.model_fields.keys()
        for field in invoice_fields:
            value = getattr(input_data, field, None)
            if value is not None:
                setattr(invoice, field, value)

        session.add(invoice)
        await session.commit()
        await session.refresh(invoice)

        return UpdateInvoiceResponse.model_validate(invoice.model_dump())

    @staticmethod
    async def delete_invoices(
        auth_user: AuthUser, input_data: DeleteInvoiceRequest
    ) -> DeleteInvoiceResponse:
        """Bulk delete invoices by IDs ensuring ownership."""

        session = auth_user.session
        ids = input_data.id
        user_id = get_user_id(auth_user)

        if not ids:
            return DeleteInvoiceResponse(
                message="No invoices specified for deletion", detail={"deleted": 0}
            )

        # Fetch invoices to verify ownership
        stmt = (
            select(Invoice)
            .where(col(Invoice.user_id) == user_id)
            .where(col(Invoice.id).in_(ids))
        )

        result = await session.exec(stmt)
        invoices: List[Invoice] = list(result.all())

        # Ownership & existence checks
        found_ids = {i.id for i in invoices}
        missing_ids = [str(i) for i in ids if i not in found_ids]

        if missing_ids:
            raise NotFoundException(
                message="Some invoices not found",
                detail="One or more invoice IDs do not exist",
                extra={"missing_ids": missing_ids},
            )

        for invoice in invoices:
            await session.delete(invoice)

        await session.commit()

        return DeleteInvoiceResponse(
            message=f"Deleted {len(invoices)} invoices successfully",
            detail={"deleted": len(invoices), "ids": [str(i.id) for i in invoices]},
        )


class InvoiceItemService:
    @staticmethod
    async def get_invoice_items(
        auth_user: AuthUser, input_params: GetInvoiceItemsRequest
    ) -> GetInvoiceItemsResponse:
        """
        Retrieve invoice items with filtering & pagination.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_params: Filtering & pagination parameters.

        Returns:
            GetInvoiceItemResponse: Paginated invoice item list.
        """

        session = auth_user.session

        # Base conditions
        conditions: List[object] = []

        # Check if ID filter is applied
        if input_params.id:
            conditions.append(InvoiceItem.id == input_params.id)

        # If ID is provided, ignore other filters and fetch by ID only
        else:
            if input_params.invoice_id:
                conditions.append(InvoiceItem.invoice_id == input_params.invoice_id)

            if input_params.product_id:
                conditions.append(InvoiceItem.product_id == input_params.product_id)

            if input_params.sale_type:
                conditions.append(InvoiceItem.sale_type == input_params.sale_type)

            if input_params.quantity is not None:
                conditions.append(InvoiceItem.quantity == input_params.quantity)

            if input_params.unit_price is not None:
                conditions.append(InvoiceItem.unit_price == input_params.unit_price)

            if input_params.retail_price is not None:
                conditions.append(InvoiceItem.retail_price == input_params.retail_price)

            if input_params.discount_percentage is not None:
                conditions.append(
                    InvoiceItem.discount_percentage == input_params.discount_percentage
                )

            if input_params.tax_rate is not None:
                conditions.append(InvoiceItem.tax_rate == input_params.tax_rate)

            if input_params.line_total is not None:
                conditions.append(InvoiceItem.line_total == input_params.line_total)

        page = input_params.page
        page_size = input_params.page_size

        # Count total
        count_stmt = select(func.count()).select_from(InvoiceItem).where(*conditions)  # type: ignore[arg-type]
        total_result = await session.exec(count_stmt)
        total = total_result.one()

        # Pagination
        stmt = (
            select(InvoiceItem).where(*conditions)  # type: ignore[arg-type]
        )

        # Apply pagination
        stmt = await apply_pagination_sql(stmt, page, page_size)

        # Apply ordering
        stmt = await apply_ordering_sql(
            stmt, InvoiceItem, input_params.order, input_params.order_by
        )

        result = await session.exec(stmt)
        rows: Sequence[InvoiceItem] = result.all()

        # Cast/validate to response models
        data: List[InvoiceItemResponse] = [
            InvoiceItemResponse.model_validate(r.model_dump()) for r in rows
        ]

        next_page = total > page * page_size
        return GetInvoiceItemsResponse(data=data, total=total, next_page=next_page)

    @staticmethod
    async def create_invoice_items(
        auth_user: AuthUser, input_data: CreateInvoiceItemsRequest
    ) -> CreateInvoiceItemsResponse:
        """
        Create multiple invoice items in bulk.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_data: List of invoice item creation data.
        """

        session = auth_user.session
        input_items = input_data.data

        invoice_items = [InvoiceItem(**item.model_dump()) for item in input_items]
        session.add_all(invoice_items)
        await session.commit()

        # Refresh all to get IDs and defaults
        for item in invoice_items:
            await session.refresh(item)

        return CreateInvoiceItemsResponse(
            data=[
                CreateInvoiceItemResponse.model_validate(item.model_dump())
                for item in invoice_items
            ]
        )

    @staticmethod
    async def update_invoice_item(
        auth_user: AuthUser, input_data: UpdateInvoiceItemRequest
    ) -> UpdateInvoiceItemResponse:
        """
        Update an existing invoice item.

        Args:
            auth_user: Authenticated user context containing DB session.
            input_data: Invoice item update data.
        """

        session = auth_user.session

        stmt = select(InvoiceItem).where(InvoiceItem.id == input_data.id)
        result = await session.exec(stmt)
        invoice_item = result.first()

        if not invoice_item:
            raise NotFoundException(
                message="Invoice item not found",
                detail=f"No invoice item found with ID {input_data.id}",
                extra={"invoice_item_id": str(input_data.id)},
            )

        invoice_item_fields = InvoiceItem.model_fields.keys()
        for field in invoice_item_fields:
            value = getattr(input_data, field, None)
            if value is not None:
                setattr(invoice_item, field, value)

        session.add(invoice_item)
        await session.commit()
        await session.refresh(invoice_item)

        return UpdateInvoiceItemResponse.model_validate(invoice_item.model_dump())

    @staticmethod
    async def update_invoice_items(
        auth_user: AuthUser, input_data: UpdateInvoiceItemsRequest
    ) -> UpdateInvoiceItemsResponse:
        """
        Bulk update multiple invoice items.

        Args:
            auth_user (AuthUser): Authenticated user context containing DB session.
            input_data (UpdateInvoiceItemsRequest): Invoice items update data.

        Raises:
            NotFoundException: If any invoice item is not found.

        Returns:
            UpdateInvoiceItemsResponse: Updated invoice items.
        """

        session = auth_user.session
        input_items = input_data.data

        item_ids: List[UUID] = [item.id for item in input_items]
        invoice_ids: List[UUID] = [
            item.invoice_id for item in input_items if item.invoice_id
        ]

        if len(invoice_ids) > 1:
            raise BadRequestException(
                message="All invoice items must belong to the same invoice for bulk update",
                detail="Provided invoice items belong to multiple invoices",
                extra={"invoice_ids": [str(i) for i in invoice_ids]},
            )

        # Fetch existing invoice items
        stmt = (
            select(InvoiceItem)
            .where(InvoiceItem.invoice_id == invoice_ids[0])
            .where(col(InvoiceItem.id).in_(item_ids))
        )
        result = await session.exec(stmt)
        existing_items: List[InvoiceItem] = list(result.all())

        # Existence checks
        found_ids = {i.id for i in existing_items}
        missing_ids = [str(i) for i in item_ids if i not in found_ids]
        if missing_ids:
            raise NotFoundException(
                message="Some invoice items not found",
                detail="One or more invoice item IDs do not exist",
                extra={"missing_ids": missing_ids},
            )

        # Create a map of existing items by ID for O(1) lookup
        existing_items_map = {item.id: item for item in existing_items}

        # Get model fields
        invoice_item_fields = InvoiceItem.model_fields.keys()

        # Update fields
        for input_item in input_items:
            existing_item = existing_items_map.get(input_item.id)
            if existing_item:
                for field in invoice_item_fields:
                    value = getattr(input_item, field, None)
                    if value is not None:
                        setattr(existing_item, field, value)
                session.add(existing_item)

        await session.commit()

        # Refresh all to get latest data
        for item in existing_items:
            await session.refresh(item)

        return UpdateInvoiceItemsResponse(
            data=[
                UpdateInvoiceItemResponse.model_validate(item.model_dump())
                for item in existing_items
            ]
        )

    @staticmethod
    async def delete_invoice_items(
        auth_user: AuthUser, input_data: DeleteInvoiceItemRequest
    ) -> DeleteInvoiceItemResponse:
        """Bulk delete invoice items by IDs."""

        session = auth_user.session
        ids = input_data.id

        if not ids:
            return DeleteInvoiceItemResponse(
                message="No invoice items specified for deletion", detail={"deleted": 0}
            )

        # Fetch invoice items
        stmt = select(InvoiceItem).where(col(InvoiceItem.id).in_(ids))

        result = await session.exec(stmt)
        invoice_items: List[InvoiceItem] = list(result.all())

        # Existence checks
        found_ids = {i.id for i in invoice_items}
        missing_ids = [str(i) for i in ids if i not in found_ids]

        if missing_ids:
            raise NotFoundException(
                message="Some invoice items not found",
                detail="One or more invoice item IDs do not exist",
                extra={"missing_ids": missing_ids},
            )

        for invoice_item in invoice_items:
            await session.delete(invoice_item)

        await session.commit()

        return DeleteInvoiceItemResponse(
            message=f"Deleted {len(invoice_items)} invoice items successfully",
            detail={
                "deleted": len(invoice_items),
                "ids": [str(i.id) for i in invoice_items],
            },
        )


# --------------------------------------------------------------------------- #
#                             Import Services                                 #
# --------------------------------------------------------------------------- #
class InvoiceImportService:
    """Utility service for parsing invoice import Excel files."""

    EXPECTED_HEADERS: List[str] = [
        "Registration No.",
        "Name",
        "Phone No.",
        "Email",
        "Registration Type",
        "Province",
        "Address",
        "Invoice Type",
        "Invoice No.",
        "Invoice Date",
        "Product Description",
        "HS Code",
        "Sale Type",
        "Rate",
        "UOM",
        "Quantity",
        "Value of Sales Excl. ST",
        "Sales Tax / FED in ST Mode",
        "Fixed / notified value or ST withheld as WH Agent",
        "Sales Tax Witheld",
        "Extra Tax",
        "Further Tax",
        "SRO No.",
        "SRO Item Sr No.",
    ]

    HEADER_KEY_MAP: Dict[str, str] = {
        "Registration No.": "registration_number",
        "Name": "name",
        "Phone No.": "phone_number",
        "Email": "email",
        "Registration Type": "registration_type",
        "Province": "province",
        "Address": "address",
        "Invoice Type": "invoice_type",
        "Invoice No.": "invoice_number",
        "Invoice Date": "invoice_date",
        "Product Description": "product_description",
        "HS Code": "hs_code",
        "Sale Type": "sale_type",
        "Rate": "rate",
        "UOM": "unit_of_measurement",
        "Quantity": "quantity",
        "Value of Sales Excl. ST": "value_of_sales_excl_st",
        "Sales Tax / FED in ST Mode": "sales_tax_fed_in_st_mode",
        "Fixed / notified value or ST withheld as WH Agent": "fixed_notified_value_or_st_withheld_as_wh_agent",
        "Sales Tax Witheld": "sales_tax_withheld",
        "Extra Tax": "extra_tax",
        "Further Tax": "further_tax",
        "SRO No.": "sro_number",
        "SRO Item Sr No.": "sro_item_serial_number",
    }

    KEY_HEADER_MAP: Dict[str, str] = {
        value: key for key, value in HEADER_KEY_MAP.items()
    }

    REQUIRED_ROW_FIELDS: List[str] = [
        "invoice_number",
        "invoice_date",
        "product_description",
        "hs_code",
        "quantity",
        "value_of_sales_excl_st",
        "name",
    ]

    BASIC_INFO_FIELDS: List[str] = [
        "registration_number",
        "name",
        "phone_number",
        "email",
        "registration_type",
        "province",
        "address",
        "invoice_type",
        "invoice_date",
    ]

    NUMERIC_FIELDS: List[str] = [
        "rate",
        "quantity",
        "value_of_sales_excl_st",
        "sales_tax_fed_in_st_mode",
        "fixed_notified_value_or_st_withheld_as_wh_agent",
        "sales_tax_withheld",
        "extra_tax",
        "further_tax",
    ]

    DATE_FIELDS: List[str] = ["invoice_date"]

    DATE_FORMATS: Tuple[str, ...] = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
    )

    EXPORT_SHEET_NAME = "Invoices"

    @classmethod
    async def export_invoices_excel(
        cls,
        auth_user: AuthUser,
        input_data: ExportInvoicesExcelRequest,
    ) -> Tuple[BytesIO, str]:
        """Export selected invoices using the same header layout as the invoice import file."""

        if not input_data.id:
            raise BadRequestException(
                message="Invoice IDs are required",
                detail="Provide at least one invoice ID to export.",
            )

        session = auth_user.session
        user_id = get_user_id(auth_user, input_data.user_id)
        invoice_ids = list(dict.fromkeys(input_data.id))

        invoice_stmt = (
            select(Invoice)
            .where(Invoice.user_id == user_id)
            .where(col(Invoice.id).in_(invoice_ids))
        )
        invoice_result = await session.exec(invoice_stmt)
        invoices = invoice_result.all()

        invoices_by_id = {invoice.id: invoice for invoice in invoices}
        missing_ids = [
            invoice_id for invoice_id in invoice_ids if invoice_id not in invoices_by_id
        ]
        if missing_ids:
            raise NotFoundException(
                message="Some invoices not found",
                detail="One or more invoice IDs do not exist.",
                extra={"missing_ids": [str(invoice_id) for invoice_id in missing_ids]},
            )

        item_stmt = select(InvoiceItem).where(
            col(InvoiceItem.invoice_id).in_(invoice_ids)
        )
        item_result = await session.exec(item_stmt)
        invoice_items = item_result.all()

        items_by_invoice_id: Dict[UUID, List[InvoiceItem]] = defaultdict(list)
        product_ids: List[UUID] = []
        for item in invoice_items:
            items_by_invoice_id[item.invoice_id].append(item)
            product_ids.append(item.product_id)

        customer_ids = list({invoice.customer_id for invoice in invoices})
        product_ids = list(dict.fromkeys(product_ids))

        customer_stmt = select(Customer).where(col(Customer.id).in_(customer_ids))
        customer_result = await session.exec(customer_stmt)
        customers = customer_result.all()
        customers_by_id = {customer.id: customer for customer in customers}

        products_by_id: Dict[UUID, Product] = {}
        if product_ids:
            product_stmt = select(Product).where(col(Product.id).in_(product_ids))
            product_result = await session.exec(product_stmt)
            products = product_result.all()
            products_by_id = {product.id: product for product in products}

        workbook = Workbook()
        worksheet = cast(Any, workbook.active)
        worksheet.title = cls.EXPORT_SHEET_NAME
        worksheet.append(cls.EXPECTED_HEADERS)

        for invoice_id in invoice_ids:
            invoice = invoices_by_id[invoice_id]
            customer = customers_by_id.get(invoice.customer_id)
            items = items_by_invoice_id.get(invoice_id, [])

            for item in items:
                product = products_by_id.get(item.product_id)
                worksheet.append(
                    cls._build_export_row(invoice, customer, item, product)
                )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        return output, f"invoices-export-{timestamp}.xlsx"

    @classmethod
    async def parse_import_file(
        cls,
        auth_user: AuthUser,
        filename: str,
        file_bytes: bytes,
    ) -> InvoiceImportParseResponse:
        """Parse an uploaded Excel file and return structured invoice data."""

        _ = getattr(auth_user.user, "id", None)  # Used for future auditing/logging.

        extension = cls._extract_extension(filename)
        if extension not in {".xlsx", ".xls"}:
            error = InvoiceImportValidationError(
                row=1,
                field="File",
                message="Invalid file type. Please upload an Excel file (.xlsx or .xls).",
                value=filename,
            )
            return InvoiceImportParseResponse(
                success=False,
                data=[],
                errors=[error],
                total_rows=0,
                valid_invoices=0,
                processed_items=0,
            )

        rows, header_errors = cls._load_rows(file_bytes, extension)
        if header_errors:
            return InvoiceImportParseResponse(
                success=False,
                data=[],
                errors=header_errors,
                total_rows=0,
                valid_invoices=0,
                processed_items=0,
            )

        total_rows = len(rows)

        cleaned_rows, row_errors = cls._clean_and_validate_rows(rows)
        if row_errors:
            return InvoiceImportParseResponse(
                success=False,
                data=[],
                errors=row_errors,
                total_rows=total_rows,
                valid_invoices=0,
                processed_items=0,
            )

        if not cleaned_rows:
            error = InvoiceImportValidationError(
                row=2,
                field="Rows",
                message="Excel file does not contain any data rows.",
                value=None,
            )
            return InvoiceImportParseResponse(
                success=False,
                data=[],
                errors=[error],
                total_rows=total_rows,
                valid_invoices=0,
                processed_items=0,
            )

        grouped_rows = cls._group_rows_by_invoice(cleaned_rows)
        group_errors = cls._validate_invoice_groups(grouped_rows)
        if group_errors:
            return InvoiceImportParseResponse(
                success=False,
                data=[],
                errors=group_errors,
                total_rows=total_rows,
                valid_invoices=0,
                processed_items=0,
            )

        parsed_invoices = cls._build_invoices(grouped_rows)
        processed_items = sum(len(invoice.items) for invoice in parsed_invoices)

        return InvoiceImportParseResponse(
            success=True,
            data=parsed_invoices,
            errors=[],
            total_rows=total_rows,
            valid_invoices=len(parsed_invoices),
            processed_items=processed_items,
        )

    @staticmethod
    def _extract_extension(filename: str) -> str:
        if not filename or "." not in filename:
            return ""
        return f".{filename.rsplit('.', 1)[-1].lower()}"

    @classmethod
    def _build_export_row(
        cls,
        invoice: Invoice,
        customer: Optional[Customer],
        item: InvoiceItem,
        product: Optional[Product],
    ) -> List[Any]:
        value_of_sales = cls._normalize_export_number(
            cls._resolve_item_value_excl_tax(item, product)
        )
        sales_tax_fed = cls._normalize_export_number(
            cls._calculate_item_tax_amount(
                cls._resolve_item_value_excl_tax(item, product), item.tax_rate
            )
        )

        return [
            cls._string_or_empty(
                customer.sales_tax_registration_number if customer else None,
                fallback=customer.national_tax_number if customer else None,
            ),
            cls._string_or_empty(customer.name if customer else None),
            cls._string_or_empty(customer.phone if customer else None),
            cls._string_or_empty(customer.email if customer else None),
            cls._enum_or_empty(customer.registration_type if customer else None),
            cls._enum_or_empty(customer.province if customer else None),
            cls._string_or_empty(customer.address if customer else None),
            cls._enum_or_empty(invoice.invoice_type),
            cls._string_or_empty(invoice.invoice_number),
            cls._format_export_date(invoice.issue_date),
            cls._string_or_empty(
                product.name if product else None,
                fallback=product.description if product else None,
            ),
            cls._string_or_empty(product.hs_code if product else None),
            cls._enum_or_empty(
                item.sale_type, fallback=product.sale_type if product else None
            ),
            cls._normalize_export_number(
                item.unit_price
                if item.unit_price is not None
                else product.unit_price
                if product
                else None
            ),
            cls._string_or_empty(product.unit_of_measurement if product else None),
            cls._normalize_export_number(item.quantity),
            value_of_sales,
            sales_tax_fed,
            cls._normalize_export_number(
                product.federal_advance_duty_payable if product else None
            ),
            cls._normalize_export_number(
                product.sales_tax_withheld if product else None
            ),
            cls._normalize_export_number(product.extra_tax if product else None),
            cls._normalize_export_number(product.further_tax if product else None),
            cls._enum_or_empty(product.sro_schedule_code if product else None),
            cls._string_or_empty(product.sro_serial_number if product else None),
        ]

    @staticmethod
    def _resolve_item_value_excl_tax(
        item: InvoiceItem,
        product: Optional[Product],
    ) -> Optional[float]:
        """
        Compute the tax-exclusive sales value for a line item from its
        price, quantity, and discount - mirroring how the invoice form
        computes it (quantity * price, less any discount).

        `InvoiceItem.line_total` is intentionally NOT used for this: for
        invoices created via the invoice form it is stored tax-INCLUSIVE
        (taxable amount + tax), so using it directly as "Value of Sales
        Excl. ST" double counts tax in the export.
        """
        sale_type_value = getattr(item.sale_type, "value", item.sale_type)
        if sale_type_value == "3rd Schedule Goods":
            price = item.retail_price
        else:
            price = (
                item.unit_price
                if item.unit_price is not None
                else product.unit_price
                if product
                else None
            )

        if price is None or item.quantity is None:
            return None

        gross = float(item.quantity) * float(price)
        discount_percentage = item.discount_percentage or 0.0
        discount_amount = gross * (float(discount_percentage) / 100.0)
        return max(gross - discount_amount, 0.0)

    @staticmethod
    def _calculate_item_tax_amount(
        line_total: Optional[float],
        tax_rate: Optional[float],
    ) -> Optional[float]:
        if line_total is None or tax_rate is None:
            return None
        return float(line_total) * (float(tax_rate) / 100.0)

    @staticmethod
    def _format_export_date(value: Optional[datetime]) -> str:
        if value is None:
            return ""
        return value.date().isoformat()

    @staticmethod
    def _enum_or_empty(value: Any, fallback: Any = None) -> str:
        resolved = value if value not in (None, "") else fallback
        if resolved is None:
            return ""
        return str(getattr(resolved, "value", resolved))

    @staticmethod
    def _string_or_empty(value: Optional[str], fallback: Optional[str] = None) -> str:
        resolved = value or fallback
        return resolved or ""

    @staticmethod
    def _normalize_export_number(value: Any) -> Any:
        if value is None:
            return ""
        try:
            return float(value)
        except (TypeError, ValueError):
            return value

    @classmethod
    def _load_rows(
        cls,
        file_bytes: bytes,
        extension: str,
    ) -> Tuple[List[Tuple[int, Dict[str, Any]]], List[InvoiceImportValidationError]]:
        rows: List[Tuple[int, Dict[str, Any]]] = []
        errors: List[InvoiceImportValidationError] = []

        try:
            if extension == ".xlsx":
                workbook = load_workbook(BytesIO(file_bytes), data_only=True)
                worksheet = cast(Any, workbook.active)
                if worksheet is None:
                    errors.append(
                        InvoiceImportValidationError(
                            row=1,
                            field="Header",
                            message="Workbook does not contain an active worksheet.",
                            value=None,
                        )
                    )
                    return rows, errors

                header_row = next(
                    worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                    None,
                )
                if not header_row:
                    errors.append(
                        InvoiceImportValidationError(
                            row=1,
                            field="Header",
                            message="Excel file is empty.",
                            value=None,
                        )
                    )
                    return rows, errors

                normalized_headers = [
                    cls._normalize_header(value) for value in header_row
                ]
                cls._validate_headers(normalized_headers, errors)
                if errors:
                    return rows, errors

                header_keys = [
                    cls.HEADER_KEY_MAP.get(header) for header in normalized_headers
                ]

                for row_index, values in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    mapped_row = cls._map_row(values, header_keys)
                    if cls._row_has_data(mapped_row):
                        rows.append((row_index, mapped_row))

            else:  # .xls
                workbook = xlrd.open_workbook(file_contents=file_bytes)
                sheet = workbook.sheet_by_index(0)
                if sheet.nrows == 0:
                    errors.append(
                        InvoiceImportValidationError(
                            row=1,
                            field="Header",
                            message="Excel file is empty.",
                            value=None,
                        )
                    )
                    return rows, errors

                header_row = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                normalized_headers = [
                    cls._normalize_header(value) for value in header_row
                ]
                cls._validate_headers(normalized_headers, errors)
                if errors:
                    return rows, errors

                header_keys = [
                    cls.HEADER_KEY_MAP.get(header) for header in normalized_headers
                ]

                for row_idx in range(1, sheet.nrows):
                    row_values = [
                        sheet.cell_value(row_idx, col) for col in range(sheet.ncols)
                    ]
                    mapped_row = cls._map_row(row_values, header_keys)
                    if cls._row_has_data(mapped_row):
                        rows.append((row_idx + 1, mapped_row))

        except Exception as exc:  # pragma: no cover - defensive guard
            errors.append(
                InvoiceImportValidationError(
                    row=1,
                    field="File",
                    message=f"Failed to read Excel file: {exc}",
                    value=None,
                )
            )
            rows.clear()

        return rows, errors

    @classmethod
    def _map_row(
        cls,
        values: Sequence[Any],
        header_keys: List[Optional[str]],
    ) -> Dict[str, Any]:
        mapped_row: Dict[str, Any] = {}
        for idx, value in enumerate(values):
            if idx >= len(header_keys):
                continue
            key = header_keys[idx]
            if key:
                mapped_row[key] = value
        return mapped_row

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _row_has_data(row: Dict[str, Any]) -> bool:
        for value in row.values():
            if value is None:
                continue
            if isinstance(value, str) and value.strip():
                return True
            if isinstance(value, (int, float)) and not (
                isinstance(value, float) and math.isnan(value)
            ):
                return True
        return False

    @classmethod
    def _validate_headers(
        cls,
        normalized_headers: List[str],
        errors: List[InvoiceImportValidationError],
    ) -> None:
        missing = [
            header
            for header in cls.EXPECTED_HEADERS
            if header not in normalized_headers
        ]
        for header in missing:
            errors.append(
                InvoiceImportValidationError(
                    row=1,
                    field="Header",
                    message=f"Missing required column '{header}'.",
                    value=header,
                )
            )

    @classmethod
    def _clean_and_validate_rows(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
    ) -> Tuple[List[Tuple[int, Dict[str, Any]]], List[InvoiceImportValidationError]]:
        cleaned_rows: List[Tuple[int, Dict[str, Any]]] = []
        errors: List[InvoiceImportValidationError] = []

        for row_index, row in rows:
            cleaned = cls._sanitize_row(row)
            if cls._row_is_empty(cleaned):
                continue

            row_errors = cls._validate_row(row_index, row, cleaned)
            if row_errors:
                errors.extend(row_errors)
                continue

            cleaned_rows.append((row_index, cleaned))

        return cleaned_rows, errors

    @classmethod
    def _sanitize_row(cls, row: Dict[str, Any]) -> Dict[str, Any]:
        cleaned: Dict[str, Any] = {}
        for key, value in row.items():
            if key in cls.NUMERIC_FIELDS:
                cleaned[key] = cls._parse_numeric(value)
            elif key in cls.DATE_FIELDS:
                cleaned[key] = cls._parse_date(value)
            else:
                cleaned[key] = cls._normalize_string(value)
        return cleaned

    @staticmethod
    def _row_is_empty(row: Dict[str, Any]) -> bool:
        return not any(value is not None for value in row.values())

    @classmethod
    def _validate_row(
        cls,
        row_index: int,
        raw_row: Dict[str, Any],
        cleaned_row: Dict[str, Any],
    ) -> List[InvoiceImportValidationError]:
        errors: List[InvoiceImportValidationError] = []

        for field in cls.REQUIRED_ROW_FIELDS:
            header = cls.KEY_HEADER_MAP.get(field, field)
            cleaned_value = cleaned_row.get(field)
            raw_value = raw_row.get(field)
            if cleaned_value is None:
                if cls._has_value(raw_value):
                    if field in cls.NUMERIC_FIELDS:
                        message = f"{header} must be a valid non-negative number."
                    elif field in cls.DATE_FIELDS:
                        message = f"{header} must be a valid date in YYYY-MM-DD format or an Excel serial date."
                    else:
                        message = f"{header} is required."
                else:
                    message = f"{header} is required."

                errors.append(
                    InvoiceImportValidationError(
                        row=row_index,
                        field=header,
                        message=message,
                        value=cls._stringify_value(raw_value),
                    )
                )

        for field in cls.NUMERIC_FIELDS:
            if field in cls.REQUIRED_ROW_FIELDS:
                continue
            raw_value = raw_row.get(field)
            if not cls._has_value(raw_value):
                continue
            parsed_value = cleaned_row.get(field)
            if parsed_value is None or (
                isinstance(parsed_value, float) and parsed_value < 0
            ):
                header = cls.KEY_HEADER_MAP.get(field, field)
                errors.append(
                    InvoiceImportValidationError(
                        row=row_index,
                        field=header,
                        message=f"{header} must be a valid non-negative number.",
                        value=cls._stringify_value(raw_value),
                    )
                )

        for field in cls.DATE_FIELDS:
            if field in cls.REQUIRED_ROW_FIELDS:
                continue
            raw_value = raw_row.get(field)
            if not cls._has_value(raw_value):
                continue
            if cleaned_row.get(field) is None:
                header = cls.KEY_HEADER_MAP.get(field, field)
                errors.append(
                    InvoiceImportValidationError(
                        row=row_index,
                        field=header,
                        message=(
                            f"{header} must be a valid date in YYYY-MM-DD format or an Excel serial date."
                        ),
                        value=cls._stringify_value(raw_value),
                    )
                )

        return errors

    @staticmethod
    def _has_value(value: object) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        if isinstance(value, (int, float)):
            return not (isinstance(value, float) and math.isnan(value))
        return True

    @staticmethod
    def _stringify_value(value: object) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        return str(value)

    @staticmethod
    def _normalize_string(value: object) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        if isinstance(value, float):
            if math.isnan(value):
                return None
            if float(value).is_integer():
                return str(int(value))
            return str(value)
        return str(value)

    @classmethod
    def _parse_numeric(cls, value: object) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return None
            return float(value)
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    @classmethod
    def _parse_date(cls, value: object) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return None
            base_date = datetime(1899, 12, 30)
            try:
                return base_date + timedelta(days=float(value))
            except (ValueError, OverflowError):
                return None
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return None
            for fmt in cls.DATE_FORMATS:
                try:
                    return datetime.strptime(stripped, fmt)
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(stripped)
            except ValueError:
                return None
        return None

    @classmethod
    def _group_rows_by_invoice(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
    ) -> Dict[str, List[Tuple[int, Dict[str, Any]]]]:
        grouped: Dict[str, List[Tuple[int, Dict[str, Any]]]] = defaultdict(list)
        for row_index, row in rows:
            invoice_number = row.get("invoice_number")
            if isinstance(invoice_number, str) and invoice_number:
                grouped[invoice_number].append((row_index, row))
        return grouped

    @classmethod
    def _validate_invoice_groups(
        cls,
        grouped_rows: Dict[str, List[Tuple[int, Dict[str, Any]]]],
    ) -> List[InvoiceImportValidationError]:
        errors: List[InvoiceImportValidationError] = []

        for invoice_number, rows in grouped_rows.items():
            for field in cls.BASIC_INFO_FIELDS:
                header = cls.KEY_HEADER_MAP.get(field, field)
                populated_values = [
                    (row_index, rows_dict.get(field))
                    for row_index, rows_dict in rows
                    if rows_dict.get(field) is not None
                ]

                if not populated_values:
                    errors.append(
                        InvoiceImportValidationError(
                            row=rows[0][0],
                            field=header,
                            message=(
                                f"{header} is missing for invoice {invoice_number}. Provide it in at least one row."
                            ),
                            value=None,
                        )
                    )
                    continue

                normalized = {
                    cls._normalize_for_comparison(value)
                    for _, value in populated_values
                }

                if len(normalized) > 1:
                    conflict_rows = ", ".join(str(idx) for idx, _ in populated_values)
                    errors.append(
                        InvoiceImportValidationError(
                            row=populated_values[0][0],
                            field=header,
                            message=(
                                f"{header} contains conflicting values for invoice {invoice_number} (rows {conflict_rows})."
                            ),
                            value=", ".join(
                                sorted(
                                    {
                                        cls._stringify_value(val) or ""
                                        for _, val in populated_values
                                    }
                                )
                            ),
                        )
                    )

        return errors

    @staticmethod
    def _normalize_for_comparison(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value).strip().lower()

    @classmethod
    def _build_invoices(
        cls,
        grouped_rows: Dict[str, List[Tuple[int, Dict[str, Any]]]],
    ) -> List[InvoiceImportParsedInvoice]:
        invoices: List[InvoiceImportParsedInvoice] = []

        for invoice_number, rows in grouped_rows.items():
            items = cls._build_items(rows)
            totals = cls._calculate_totals(items)

            invoice = InvoiceImportParsedInvoice(
                invoice_number=invoice_number,
                issue_date=cls._choose_first_datetime(rows, "invoice_date"),
                invoice_type=cls._choose_first_str(rows, "invoice_type")
                or "Sale Invoice",
                registration_number=cls._choose_first_str(rows, "registration_number"),
                buyer_name=cls._choose_first_str(rows, "name"),
                phone_number=cls._choose_first_str(rows, "phone_number"),
                email=cls._choose_first_str(rows, "email"),
                buyer_registration_type=RegistrationTypeEnum(
                    cls._choose_first_str(rows, "registration_type")
                ),
                province=ProvinceEnum(cls._choose_first_str(rows, "province")),
                address=cls._choose_first_str(rows, "address"),
                subtotal=totals["subtotal"],
                discount_amount=0.0,
                tax_amount=totals["tax_amount"],
                total_amount=totals["total_amount"],
                extra_tax_amount=totals["extra_tax"],
                further_tax_amount=totals["further_tax"],
                federal_advance_duty_payable_amount=totals["fad_payable"],
                fbr_validated=False,
                fbr_status=InvoiceFBRStatusEnum.PENDING,
                status=InvoiceStatusEnum.DRAFT,
                items=items,
            )

            invoices.append(invoice)

        return invoices

    @classmethod
    def _build_items(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
    ) -> List[InvoiceImportParsedInvoiceItem]:
        items: List[InvoiceImportParsedInvoiceItem] = []

        for _, row in rows:
            value_of_sales = cls._safe_float(row.get("value_of_sales_excl_st"))
            tax_value = cls._safe_float(row.get("sales_tax_fed_in_st_mode"))
            tax_rate = None
            if value_of_sales and value_of_sales > 0 and tax_value:
                tax_rate = (tax_value / value_of_sales) * 100

            item = InvoiceImportParsedInvoiceItem(
                sale_type=SaleTypeEnum(cls._coerce_str(row.get("sale_type"))),
                quantity=cls._safe_int(row.get("quantity")),
                unit_price=cls._safe_float(row.get("rate")),
                discount_percentage=0.0,
                tax_rate=tax_rate,
                line_total=value_of_sales,
                product_description=cls._coerce_str(row.get("product_description")),
                hs_code=cls._coerce_str(row.get("hs_code")),
                unit_of_measurement=cls._coerce_str(row.get("unit_of_measurement")),
                value_of_sales_excl_st=value_of_sales,
                sales_tax_fed_in_st_mode=tax_value,
                fixed_notified_value_or_st_withheld_as_wh_agent=cls._safe_float(
                    row.get("fixed_notified_value_or_st_withheld_as_wh_agent")
                ),
                sales_tax_withheld=cls._safe_float(row.get("sales_tax_withheld")),
                extra_tax=cls._safe_float(row.get("extra_tax")),
                further_tax=cls._safe_float(row.get("further_tax")),
                sro_number=cls._coerce_str(row.get("sro_number")),
                sro_item_serial_number=cls._coerce_str(
                    row.get("sro_item_serial_number")
                ),
            )

            items.append(item)

        return items

    @staticmethod
    def _coerce_str(value: Any) -> Optional[str]:
        return InvoiceImportService._normalize_string(value)

    @staticmethod
    def _safe_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return None
            return float(value)
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    @staticmethod
    def _safe_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if isinstance(value, float) and math.isnan(value):
                return None
            if value.is_integer():
                return int(value)
            return None
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            try:
                float_value = float(stripped)
                if float_value.is_integer():
                    return int(float_value)
                return None
            except ValueError:
                return None
        return None

    @classmethod
    def _choose_first_str(
        cls,
        rows: List[Tuple[int, Dict[str, Any]]],
        key: str,
    ) -> Optional[str]:
        for _, row in rows:
            value = row.get(key)
            coerced = cls._coerce_str(value)
            if coerced is not None:
                return coerced
        return None

    @staticmethod
    def _choose_first_datetime(
        rows: List[Tuple[int, Dict[str, Any]]],
        key: str,
    ) -> Optional[datetime]:
        for _, row in rows:
            value = row.get(key)
            if isinstance(value, datetime):
                return value
        return None

    @staticmethod
    def _calculate_totals(
        items: List[InvoiceImportParsedInvoiceItem],
    ) -> Dict[str, float]:
        subtotal = sum(item.value_of_sales_excl_st or 0.0 for item in items)
        fed_tax = sum(item.sales_tax_fed_in_st_mode or 0.0 for item in items)
        extra_tax = sum(item.extra_tax or 0.0 for item in items)
        further_tax = sum(item.further_tax or 0.0 for item in items)
        fad_payable = sum(
            item.fixed_notified_value_or_st_withheld_as_wh_agent or 0.0
            for item in items
        )
        sales_tax_withheld = sum(item.sales_tax_withheld or 0.0 for item in items)

        tax_amount = fed_tax + sales_tax_withheld
        total_amount = (
            subtotal
            + fed_tax
            + extra_tax
            + further_tax
            + fad_payable
            + sales_tax_withheld
        )

        return {
            "subtotal": subtotal,
            "tax_amount": tax_amount,
            "total_amount": total_amount,
            "extra_tax": extra_tax,
            "further_tax": further_tax,
            "fad_payable": fad_payable,
        }

    @classmethod
    async def submit_imported_invoices(
        cls,
        auth_user: AuthUser,
        input_data: InvoiceImportSubmitRequest,
    ) -> InvoiceImportSubmitResponse:
        """Persist parsed invoices, creating dependent records as required."""

        if not input_data.invoices:
            return InvoiceImportSubmitResponse(success=True, invoices=[])

        results: List[InvoiceImportSubmitInvoiceResult] = []

        for index, invoice_input in enumerate(input_data.invoices, start=1):
            if not invoice_input.items:
                raise BadRequestException(
                    message="Invoice items are required",
                    detail="Each submitted invoice must include at least one item.",
                    extra={
                        "invoice_index": index,
                        "invoice_number": invoice_input.invoice_number,
                    },
                )

            user_id = get_user_id(auth_user, invoice_input.user_id)
            customer_id = await cls._resolve_customer(
                auth_user, user_id, invoice_input, index
            )
            invoice_request = cls._prepare_invoice_request(
                invoice_input, customer_id, user_id, index
            )

            created_invoice = await InvoiceService.create_invoice(
                auth_user, invoice_request
            )

            created_items = await cls._create_invoice_items(
                auth_user,
                invoice_input,
                created_invoice.id,
                user_id,
                index,
            )

            results.append(
                InvoiceImportSubmitInvoiceResult(
                    invoice=created_invoice,
                    items=created_items,
                )
            )

        return InvoiceImportSubmitResponse(success=True, invoices=results)

    @classmethod
    async def _resolve_customer(
        cls,
        auth_user: AuthUser,
        user_id: UUID,
        invoice_input: InvoiceImportSubmitInvoice,
        invoice_index: int,
    ) -> UUID:
        if invoice_input.customer_id:
            return invoice_input.customer_id

        customer_payload = invoice_input.customer
        if customer_payload is None:
            name = invoice_input.buyer_name or invoice_input.registration_number
            if not name:
                raise BadRequestException(
                    message="Customer information missing",
                    detail="Provide a customer_id or customer payload for each invoice.",
                    extra={
                        "invoice_index": invoice_index,
                        "invoice_number": invoice_input.invoice_number,
                    },
                )

            customer_payload = CreateCustomerRequest(
                user_id=user_id,
                name=name,
                email=invoice_input.email,
                phone=invoice_input.phone_number,
                address=invoice_input.address,
                province=ProvinceEnum(invoice_input.province)
                if invoice_input.province
                else None,
                registration_type=RegistrationTypeEnum(
                    invoice_input.buyer_registration_type
                )
                if invoice_input.buyer_registration_type
                else None,
                national_tax_number=invoice_input.registration_number,
                sales_tax_registration_number=invoice_input.registration_number,
            )
        else:
            if not customer_payload.name:
                fallback_name = (
                    invoice_input.buyer_name or invoice_input.registration_number
                )
                if not fallback_name:
                    raise BadRequestException(
                        message="Customer name missing",
                        detail="Customer payload must include a name field.",
                        extra={
                            "invoice_index": invoice_index,
                            "invoice_number": invoice_input.invoice_number,
                        },
                    )
                customer_payload.name = fallback_name

            customer_payload.user_id = customer_payload.user_id or user_id

            if not customer_payload.email and invoice_input.email:
                customer_payload.email = invoice_input.email
            if not customer_payload.phone and invoice_input.phone_number:
                customer_payload.phone = invoice_input.phone_number
            if not customer_payload.address and invoice_input.address:
                customer_payload.address = invoice_input.address
            if not customer_payload.province and invoice_input.province:
                customer_payload.province = (
                    ProvinceEnum(invoice_input.province)
                    if invoice_input.province
                    else None
                )
            if (
                not customer_payload.registration_type
                and invoice_input.buyer_registration_type
            ):
                customer_payload.registration_type = (
                    RegistrationTypeEnum(invoice_input.buyer_registration_type)
                    if invoice_input.buyer_registration_type
                    else None
                )
            if (
                not customer_payload.national_tax_number
                and invoice_input.registration_number
            ):
                customer_payload.national_tax_number = invoice_input.registration_number
            if (
                not customer_payload.sales_tax_registration_number
                and invoice_input.registration_number
            ):
                customer_payload.sales_tax_registration_number = (
                    invoice_input.registration_number
                )

        created_customer = await CustomerService.create_customer(
            auth_user, customer_payload
        )
        return created_customer.id

    @classmethod
    def _prepare_invoice_request(
        cls,
        invoice_input: InvoiceImportSubmitInvoice,
        customer_id: UUID,
        user_id: UUID,
        invoice_index: int,
    ) -> CreateInvoiceRequest:
        invoice_number = invoice_input.invoice_number
        if not invoice_number:
            raise BadRequestException(
                message="Invoice number missing",
                detail="Each submitted invoice must include an invoice_number.",
                extra={"invoice_index": invoice_index},
            )

        issue_date = invoice_input.issue_date or datetime.now(tz=timezone.utc)
        items_for_total = cast(
            List[InvoiceImportParsedInvoiceItem], invoice_input.items
        )
        totals = cls._calculate_totals(items_for_total)

        subtotal = (
            float(invoice_input.subtotal)
            if invoice_input.subtotal is not None
            else float(totals["subtotal"])
        )
        discount_amount = (
            float(invoice_input.discount_amount)
            if invoice_input.discount_amount is not None
            else 0.0
        )
        tax_amount = (
            float(invoice_input.tax_amount)
            if invoice_input.tax_amount is not None
            else float(totals["tax_amount"])
        )
        total_amount = (
            float(invoice_input.total_amount)
            if invoice_input.total_amount is not None
            else float(totals["total_amount"])
        )
        extra_tax_amount = (
            float(invoice_input.extra_tax_amount)
            if invoice_input.extra_tax_amount is not None
            else float(totals["extra_tax"])
        )
        further_tax_amount = (
            float(invoice_input.further_tax_amount)
            if invoice_input.further_tax_amount is not None
            else float(totals["further_tax"])
        )
        fad_amount = (
            float(invoice_input.federal_advance_duty_payable_amount)
            if invoice_input.federal_advance_duty_payable_amount is not None
            else float(totals["fad_payable"])
        )

        subtotal = max(subtotal, 0.0)
        discount_amount = max(discount_amount, 0.0)
        tax_amount = max(tax_amount, 0.0)
        total_amount = max(total_amount, 0.0)
        extra_tax_amount = max(extra_tax_amount, 0.0)
        further_tax_amount = max(further_tax_amount, 0.0)
        fad_amount = max(fad_amount, 0.0)

        return CreateInvoiceRequest(
            user_id=user_id,
            customer_id=customer_id,
            invoice_number=invoice_number,
            issue_date=issue_date,
            due_date=invoice_input.due_date,
            subtotal=subtotal,
            discount_amount=discount_amount,
            tax_amount=tax_amount,
            total_amount=total_amount,
            status=InvoiceStatusEnum(invoice_input.status or "draft"),
            notes=invoice_input.notes,
            invoice_type=(
                InvoiceTypeEnum(invoice_input.invoice_type or "Sale Invoice")
            ),
            extra_tax_amount=extra_tax_amount,
            further_tax_amount=further_tax_amount,
            federal_advance_duty_payable_amount=fad_amount,
            fbr_validated=invoice_input.fbr_validated
            if invoice_input.fbr_validated is not None
            else False,
            fbr_status=invoice_input.fbr_status or InvoiceFBRStatusEnum.PENDING,
            fbr_reference=invoice_input.fbr_reference,
        )

    @classmethod
    async def _create_invoice_items(
        cls,
        auth_user: AuthUser,
        invoice_input: InvoiceImportSubmitInvoice,
        invoice_id: UUID,
        user_id: UUID,
        invoice_index: int,
    ) -> List[InvoiceItemResponse]:
        item_requests: List[InvoiceItemBase] = []

        for item_index, item in enumerate(invoice_input.items, start=1):
            quantity = cls._resolve_quantity(item)
            line_total = cls._resolve_line_total(item, quantity)
            unit_price = cls._resolve_unit_price(item, quantity, line_total)
            tax_rate = cls._resolve_tax_rate(item, line_total)

            quantity = max(quantity, 0)
            line_total = max(line_total, 0.0)
            tax_rate = max(tax_rate, 0.0)
            if unit_price is not None:
                unit_price = max(unit_price, 0.0)

            product_id = await cls._ensure_product(
                auth_user,
                user_id,
                item,
                unit_price=unit_price if unit_price is not None else 0.0,
                tax_rate=tax_rate,
                invoice_index=invoice_index,
                item_index=item_index,
                invoice_number=invoice_input.invoice_number,
            )

            item_requests.append(
                InvoiceItemBase(
                    invoice_id=invoice_id,
                    product_id=product_id,
                    sale_type=item.sale_type,
                    quantity=quantity,
                    unit_price=unit_price,
                    discount_percentage=float(item.discount_percentage or 0.0),
                    tax_rate=tax_rate,
                    line_total=line_total,
                )
            )

        if not item_requests:
            return []

        response = await InvoiceItemService.create_invoice_items(
            auth_user, CreateInvoiceItemsRequest(data=item_requests)
        )

        return response.data

    @classmethod
    async def _ensure_product(
        cls,
        auth_user: AuthUser,
        user_id: UUID,
        item: InvoiceImportSubmitInvoiceItem,
        *,
        unit_price: float,
        tax_rate: float,
        invoice_index: int,
        item_index: int,
        invoice_number: Optional[str],
    ) -> UUID:
        if item.product_id:
            return item.product_id

        payload: Dict[str, Any] = (
            item.product.model_dump(exclude_unset=True) if item.product else {}
        )

        payload["user_id"] = payload.get("user_id") or user_id

        name_candidate = (
            payload.get("name")
            or item.product_description
            or (item.hs_code or "Imported product")
        )
        payload["name"] = name_candidate.strip() or "Imported product"

        payload["description"] = payload.get("description") or item.product_description
        payload["hs_code"] = payload.get("hs_code") or item.hs_code or ""
        payload["unit_of_measurement"] = (
            payload.get("unit_of_measurement") or item.unit_of_measurement or "pcs"
        )

        if payload.get("unit_price") is None:
            payload["unit_price"] = max(unit_price, 0.0)

        if not payload.get("tax_rate") and tax_rate:
            payload["tax_rate"] = f"{tax_rate:.2f}"

        if (
            payload.get("sales_tax_applicable") is None
            and item.sales_tax_fed_in_st_mode is not None
        ):
            payload["sales_tax_applicable"] = item.sales_tax_fed_in_st_mode
        if (
            payload.get("sales_tax_withheld") is None
            and item.sales_tax_withheld is not None
        ):
            payload["sales_tax_withheld"] = item.sales_tax_withheld
        if payload.get("extra_tax") is None and item.extra_tax is not None:
            payload["extra_tax"] = item.extra_tax
        if payload.get("further_tax") is None and item.further_tax is not None:
            payload["further_tax"] = item.further_tax
        if (
            payload.get("federal_advance_duty_payable") is None
            and item.fixed_notified_value_or_st_withheld_as_wh_agent is not None
        ):
            payload["federal_advance_duty_payable"] = (
                item.fixed_notified_value_or_st_withheld_as_wh_agent
            )
        if payload.get("sro_schedule_code") is None and item.sro_number is not None:
            payload["sro_schedule_code"] = item.sro_number
        if (
            payload.get("sro_serial_number") is None
            and item.sro_item_serial_number is not None
        ):
            payload["sro_serial_number"] = item.sro_item_serial_number

        if not payload["name"]:
            raise BadRequestException(
                message="Product name missing",
                detail="Unable to determine a product name from the imported data.",
                extra={
                    "invoice_index": invoice_index,
                    "item_index": item_index,
                    "invoice_number": invoice_number,
                },
            )

        product_request = CreateProductRequest(**payload)
        product_response = await ProductService.create_product(
            auth_user, product_request
        )
        return product_response.id

    @staticmethod
    def _resolve_quantity(item: InvoiceImportSubmitInvoiceItem) -> int:
        if item.quantity is not None:
            return int(item.quantity)
        return 1

    @staticmethod
    def _resolve_line_total(
        item: InvoiceImportSubmitInvoiceItem, quantity: int
    ) -> float:
        if item.line_total is not None:
            return float(item.line_total)
        if item.value_of_sales_excl_st is not None:
            return float(item.value_of_sales_excl_st)
        if item.unit_price is not None and quantity:
            return float(item.unit_price) * quantity
        return 0.0

    @staticmethod
    def _resolve_unit_price(
        item: InvoiceImportSubmitInvoiceItem,
        quantity: float,
        line_total: float,
    ) -> Optional[float]:
        if item.unit_price is not None:
            return float(item.unit_price)
        if quantity:
            return float(line_total) / float(quantity)
        if line_total:
            return float(line_total)
        return None

    @staticmethod
    def _resolve_tax_rate(
        item: InvoiceImportSubmitInvoiceItem,
        line_total: float,
    ) -> float:
        if item.tax_rate is not None:
            return float(item.tax_rate)
        if item.sales_tax_fed_in_st_mode and line_total:
            base = float(line_total) if line_total else 0.0
            if base:
                return float(item.sales_tax_fed_in_st_mode) / base * 100
        if item.product and item.product.tax_rate:
            try:
                return float(item.product.tax_rate)
            except (TypeError, ValueError):
                return 0.0
        return 0.0